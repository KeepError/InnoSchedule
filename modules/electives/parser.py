"""
This module contains all necessary functions to parse electives schedule.
"""
import calendar
from datetime import datetime
from typing import List, Union, Dict
from sqlalchemy.exc import SQLAlchemyError

import requests
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from modules.electives import permanent, controller
from modules.electives.classes import Elective, ElectiveLesson


def parse_new_electives_timetable():
    """
    Downloads new electives timetable from the link, and tries to parse it,
    putting parsed to DB. Use wisely! Users are un-enrolled when it is called
    """
    # Download
    new_schedule = requests.get(permanent.ELECTIVE_SCHEDULE_LINK)
    with open(permanent.ELECTIVE_FILE_PATH, 'wb') as f:
        f.write(new_schedule.content)
    # Delete all electives
    controller.delete_electives()
    wb = load_workbook(permanent.ELECTIVE_FILE_PATH, read_only=True)
    for sheet in wb:
        # Parse sheet
        electives = find_electives(sheet)
        lessons = find_lessons(sheet)
        # Save to db
        for elective in electives:
            # No point in saving elective w/o lessons to SCHEDULE bot db
            if elective.acronym in lessons:
                try:
                    controller.add_elective(elective)
                    child_lessons = lessons[elective.acronym]
                    for lesson in child_lessons:
                        controller.add_elective_lesson(elective.id, lesson)
                except SQLAlchemyError:
                    controller.delete_elective(elective.id)


def find_electives(sheet: Worksheet) -> List[Elective]:
    """
    Finds definition of all electives on given sheet
    """

    def strip_none(string: Union[str, None]) -> str:
        """
        Call str.strip(), if string is supplied as argument
        If argument is None, return empty string
        """
        if string is None:
            return ""
        return string.strip()

    def first_nonempty_row() -> int:
        columns = permanent.ELECTIVE_DEFINITION_COLUMNS.values()
        row = 2
        while True:
            values = [sheet.cell(row, col).value for col in columns]
            values = [strip_none(val) for val in values]
            if all(len(part) != 0 for part in values):
                return row
            row += 1

    group = sheet.title
    result = []
    electives_remain = True
    cursor_row = first_nonempty_row()
    while electives_remain:
        number = sheet.cell(cursor_row, permanent.NUMBER_COLUMN).value
        name = sheet.cell(cursor_row, permanent.ELECTIVE_DEFINITION_COLUMNS["name"]).value
        teacher = sheet.cell(cursor_row,
                             permanent.ELECTIVE_DEFINITION_COLUMNS["teacher"]).value
        acronym = sheet.cell(cursor_row,
                             permanent.ELECTIVE_DEFINITION_COLUMNS["acronym"]).value
        data = [strip_none(i) for i in [name, teacher, acronym]]
        if all(len(part) != 0 for part in data):
            result.append(Elective(*data[:3], group))
            cursor_row += 1
        elif number:
            # master courses have two categories separated by an empty line
            cursor_row += 1
        else:
            electives_remain = False
    return result


def find_lessons(sheet: Worksheet) -> Dict[str, List[ElectiveLesson]]:
    """Find lessons on the sheet, save them to Map, with course acronym as key,
    and lesson as value"""
    month_names = [calendar.month_name[i] for i in range(1, 12 + 1)]
    ignore_list = ['Block']
    # First column contains definition of timeslots, so skip it
    start_col = 2
    result: Dict[str, List[ElectiveLesson]] = {}
    for column in range(start_col, start_col + permanent.DAYS_IN_WEEK):
        # First row contains names of days of a week, so skip it
        current_month = None
        current_day = None
        for row in range(2, permanent.MAX_TIMESLOTS_IN_DAY * permanent.MAX_WEEKS):
            value: str = sheet.cell(row, column).value
            if value is None:
                # Skip empty cell
                continue
            if value.split() and value.split()[0] in ignore_list:
                continue

            if value.split() and value.split()[0] in month_names:
                # New day begins
                current_month, current_day = value.split()
                current_day = int(current_day)
            else:
                # Lesson
                if not sheet.cell(row, 1).value:
                    continue
                start_time: str = sheet.cell(row, 1).value.split(sep='-')[0]

                date_and_time = datetime(
                    year=datetime.now().year,
                    month=list(calendar.month_name).index(current_month),
                    day=current_day,
                    hour=int(start_time.split(sep=':')[0]),
                    minute=int(start_time.split(sep=':')[1])
                )

                lines = value.split("\n")
                # Even-numbered tokens are acronyms, and odds - rooms

                # acronyms: List[str] = []
                # rooms: List[int] = []
                # for i in range(len(tokens)):
                #     if tokens[i].isupper():
                #         acronyms.append(tokens[i])
                #     if tokens[i].isnumeric():
                #         rooms.append(int(tokens[i]))

                for idx, line in enumerate(lines):
                    tokens = line.split()
                    if len(tokens) == 0:
                        continue
                    acronym = tokens[0]
                    room = tokens[1] if len(tokens) > 1 else "?"
                    lesson: ElectiveLesson = ElectiveLesson(room, date_and_time)

                    lessons: List[ElectiveLesson] = [lesson]
                    if acronym in result:
                        lessons = result[acronym] + lessons

                    result[acronym] = lessons
    return result
