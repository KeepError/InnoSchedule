"""
Tests for electives module. As I did not understand how to mock sqlalchemy db,
all tests should clean after themselves
"""
import unittest
from datetime import datetime

import requests
from openpyxl import load_workbook
from telebot.types import Message

from modules.core.source import bot
from modules.electives import controller, parser
from modules.electives.classes import User, Elective, ElectiveLesson
from modules.electives.controller import DBException
from modules.electives.parser import find_electives


class DBTests(unittest.TestCase):

    def test_save_elective(self):
        new_elective = Elective("Doing Nothing", "Gandi", "DN", "Main")
        try:
            controller.add_elective(new_elective)
            # Inserted
            self.assertTrue(controller.get_elective(new_elective.id) is not None)
            # Can not be inserted second time
            self.assertRaises(DBException, controller.add_elective, new_elective)
        finally:
            controller.delete_elective(new_elective.id)

    def test_enroll(self):
        new_elective = Elective("Doing Nothing", "Gandi", "DN", "Main")
        new_user = User(chat_id=999999)

        try:
            controller.register_user(new_user)
            controller.add_elective(new_elective)

            controller.enroll(new_elective.id, new_user.chat_id)
            self.assertEqual(controller.get_electives_of_user(new_user), [new_elective])

        finally:
            controller.delete_user(new_user.chat_id)
            controller.delete_elective(new_elective.id)

    def test_lessons_auto_deleted(self):
        """Tests that when elective is deleted, its lessons are removed as well"""
        new_elective = Elective("Doing Nothing", "Gandi", "DN", "Main")
        lessons = [ElectiveLesson(104, datetime(2021, 1, 28, 14, 00))]
        try:
            controller.add_elective(new_elective)
            controller.add_elective_lesson(elective_id=new_elective.id, lesson=lessons[0])

            amount_of_lessons = len(controller.get_lessons())
            lessons_to_remove = controller.get_lessons_of_elective(new_elective.id)
            controller.delete_elective(elective_id=new_elective.id)

            self.assertEqual(amount_of_lessons - len(lessons_to_remove),
                             len(controller.get_lessons()))
        except DBException:
            try:
                for lesson in lessons:
                    controller.delete_lesson(lesson)
                controller.delete_elective(new_elective.id)
            except DBException:
                pass


class ParserTests(unittest.TestCase):
    def test_parse_electives_list(self):
        temp_schedule = requests.get(
            "https://docs.google.com/spreadsheets/d/1V_qqAuQ8k2tLJzQbUgiODbi5-V1C6E5F/export?format=xlsx")
        with open("modules/electives/tmp.xlsx", 'wb') as f:
            f.write(temp_schedule.content)
        wb = load_workbook("modules/electives/tmp.xlsx", read_only=True)
        sheet = wb[wb.sheetnames[0]]
        electives = find_electives(sheet)
        valid_electives = [
            Elective("Reverse Engineering", "Daulet Tumbaev", "RE", "BS3"),
            Elective("Project", "Optional", "PT", "BS3"),
            Elective("Systematic Literature Review", "Mirko Farina", "SLR", "BS3"),
            Elective("Applied Software Architecture", "Anna Melekhova", "ASA", "BS3"),
            Elective("Introduction to Human Computer Interaction Design for AI", "Sebastian Denef",
                     "IHCIDAI", "BS3"),
            Elective("Distributed Systems and Middleware: Patterns and Frameworks",
                     "Paolo Ciancarini", "DSMPF", "BS3"),
            Elective("Computer Graphics in Game Development", "Ivan Belyavtsev", "CGGD", "BS3"),
            Elective("Programming in Haskell", "Nikolay Kudasov", "PH", "BS3"),
            Elective("Enterprise Programming on Javascript - Advanced", "Andrey Vlasov", "EPJA",
                     "BS3"),
            Elective("Modern Application Production", "Jean-Michel Bruel", "MAP", "BS3"),
            Elective("Consensus Theory and Concurrent Programming on a Shared Memory",
                     "Alexander Tormasov, Artem Burmyakov", "CTCPSM", "BS3"),
            Elective("Advanced Topics in Software Testing", "Mohamad Kassab", "ATST", "BS3"),
        ]
        self.assertEqual(valid_electives, electives)


def run_tests(message: Message):
    elective_suite = unittest.TestSuite()
    elective_suite.addTest(unittest.makeSuite(DBTests))
    elective_suite.addTest(unittest.makeSuite(ParserTests))
    print("count of tests: " + str(elective_suite.countTestCases()) + "\n")

    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(elective_suite)
    if result.wasSuccessful():
        bot.send_message(message.chat.id,
                         f"Successfully ran all tests ({result.testsRun} tests were run)")
    else:
        bot.send_message(message.chat.id, f"Some tests failed: \n {result.failures}")
