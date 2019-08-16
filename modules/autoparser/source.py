import re
import requests
import shutil

from openpyxl import load_workbook
import schedule

from modules.autoparser import controller
from modules.autoparser import permanent
from modules.core.permanent import DATABASE_FOLDER, DATABASE_NAME
from modules.core.source import bot
from modules.admin.permanent import ADMIN_NOTIFY_LIST

"""
Module automatically parse schedule from google sheet and modify database

Author: @Nmikriukov
"""


def attach_autoparser_module():

    def get_value(ws, row, col):
        """
        Get value of specific cell from worksheet
        If cell is part of merged cell, return value from top left single cell, where text is stored

        :param ws: worksheet from openpyxl workbook
        :param row: integer
        :param col: integer
        :return: text
        """
        # check if cell is merged
        for borders in ws.merged_cells.ranges:
            if borders.min_col <= col <= borders.max_col and borders.min_row <= row <= borders.max_row:
                return ws.cell(borders.min_row, borders.min_col).value
        # not merged cell
        return ws.cell(row, col).value

    def parse_cell(ws, row, col):
        """
        Get lesson, teacher and room from specific cell

        :param ws: worksheet from openpyxl workbook
        :param row: integer
        :param col: integer
        :return: text, text, text | if correct cell with data
                 None, None, None | if empty cell
                 -1, None, None   | if unknown data in cell
        """
        lesson = get_value(ws, row, col)
        teacher = get_value(ws, row + 1, col)
        room = get_value(ws, row + 2, col)
        if not lesson or len(lesson) < 2:
            return None, None, None  # empty cell

        # remove () brackets and strip
        if lesson:
            lesson = re.sub(r"\(.+\)", "", lesson).strip()
        if teacher:
            teacher = re.sub(r"\(.+\)", "", teacher).strip()
        if isinstance(room, str):
            room = re.sub(r"\(.+\)", "", room).strip()

        if not teacher:  # unknown teacher
            teacher = '?'
        try:
            room = int(room)
        except TypeError:
            room = -1
        except ValueError:  # unknown room
            return -1, None, None

        return lesson, teacher, room

    # specific Exception for download error
    class ScheduleDownloadError(Exception):
        pass

    def parse_new_timetable():
        """
        Download xlsx schedule from link and parse all lessons
        Stores two previous versions of databases and xlsx files
        """
        try:
            # move previous backups
            shutil.move(f"{DATABASE_FOLDER}/{permanent.DATABASE_BACKUP_1}",
                        f"{DATABASE_FOLDER}/{permanent.DATABASE_BACKUP_2}")
            shutil.move(f"{DATABASE_FOLDER}/{permanent.SCHEDULE_BACKUP_1}",
                        f"{DATABASE_FOLDER}/{permanent.SCHEDULE_BACKUP_2}")
        except FileNotFoundError:
            pass
        compare_with_prev = True  # compare with previous version of database if such is found
        try:
            # make new backup
            shutil.copy(f"{DATABASE_FOLDER}/{DATABASE_NAME}",
                        f"{DATABASE_FOLDER}/{permanent.DATABASE_BACKUP_1}")
            shutil.move(f"{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}",
                        f"{DATABASE_FOLDER}/{permanent.SCHEDULE_BACKUP_1}")
        except FileNotFoundError:
            compare_with_prev = False

        # download new schedule from google sheet
        new_schedule = requests.get(permanent.SCHEDULE_DOWNLOAD_LINK)
        with open(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}', 'wb') as f:
            f.write(new_schedule.content)

        try:
            # check download is ok
            schedule_size = shutil.os.path.getsize(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}')
            if schedule_size < permanent.SCHEDULE_MIN_SIZE_BYTES:
                raise ScheduleDownloadError
        except (FileNotFoundError, ScheduleDownloadError):
            # send error notification to admins
            for admin in ADMIN_NOTIFY_LIST:
                bot.send_message(admin, permanent.MESSAGE_ERROR_NOTIFY)
            return

        # delete all lessons because new ones will be parsed
        controller.delete_all_lessons()

        # open workbook
        wb = load_workbook(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_NAME}')
        ws = wb[wb.sheetnames[0]]

        # open workbook from backup
        wb_old, ws_old = None, None
        if compare_with_prev:
            wb_old = load_workbook(f'{DATABASE_FOLDER}/{permanent.SCHEDULE_BACKUP_1}')
            ws_old = wb_old[wb_old.sheetnames[0]]

        # iterate over each cell
        col = 2
        while col <= permanent.SCHEDULE_LAST_COLUMN:
            course_group = get_value(ws, 2, col)

            if course_group[:3] == "B19":
                course_group = course_group[:6] + ('-a' if col % 2 == 0 else '-b')

            cur_weekday = 0
            row = 4
            while row <= permanent.SCHEDULE_LAST_ROW:
                first_col_value = get_value(ws, row, 1)  # time or weekday
                if isinstance(first_col_value, str) and first_col_value.upper() in permanent.WEEKDAYS:
                    cur_weekday += 1
                    row += 1
                    continue

                cell_new = parse_cell(ws, row, col)
                if not cell_new[0]:
                    row += 3
                    continue
                if cell_new[0] == -1:
                    # send error notification to admins
                    for admin in ADMIN_NOTIFY_LIST:
                        bot.send_message(admin, f"{permanent.MESSAGE_ERROR_PARSE_SYNTAX} row={row} col={col}")
                    row += 3
                    continue

                subject, teacher, room = cell_new[0], cell_new[1], cell_new[2]
                # extract time
                time_splitted = first_col_value.split('-')
                start_time, end_time = time_splitted[0], time_splitted[1]

                if compare_with_prev:
                    # compare new cell with old one
                    cell_old = parse_cell(ws_old, row, col)
                    if cell_new != cell_old:
                        subject_old, teacher_old, room_old = cell_old[0], cell_old[1], cell_old[2]
                        for admin in ADMIN_NOTIFY_LIST:
                            # send changes to admin
                            bot.send_message(admin, f"{course_group} {first_col_value} changed:\n"
                                                    f"Was {subject_old}, {teacher_old}, {room_old}\n"
                                                    f"Now {subject}, {teacher}, {room}\n")

                # insert new lesson to database
                controller.insert_lesson(course_group, subject, teacher, cur_weekday, start_time, end_time, room)
                row += 3
            col += 1

        # add special lessons here manually
        # controller.insert_lesson("B17-03", "SQL injections", "Nikolai Mikriukov", 0, "13:37", "15:00", 108)

    # open parse function to other modules
    attach_autoparser_module.parse_schedule_func = parse_new_timetable
    # add parse function call to schedule on each day
    schedule.every().day.at(permanent.ADMIN_NOTIFY_TIME).do(parse_new_timetable)
